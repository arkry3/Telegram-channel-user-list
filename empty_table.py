from openpyxl import load_workbook
def empty_table(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    wb.save(file_path)